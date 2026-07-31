from __future__ import annotations

import ast
import hashlib
import json
import re
from copy import copy
from pathlib import Path
from types import SimpleNamespace

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

ROOT = Path(__file__).parents[1]
NOTEBOOKS = [
    ROOT / "Audio_Transcription_Colab.ipynb",
    ROOT / "OCR_HTR_Colab.ipynb",
    ROOT / "Summary_Colab.ipynb",
]


def load_notebook(path):
    return json.loads(path.read_text(encoding="utf-8"))


def notebook_code(notebook):
    return "\n\n".join(
        "".join(cell["source"])
        for cell in notebook["cells"]
        if cell["cell_type"] == "code"
    )


def parse_notebook_code(path):
    notebook = load_notebook(path)
    for index, cell in enumerate(notebook["cells"]):
        if cell["cell_type"] != "code":
            continue
        lines = [
            line
            for line in "".join(cell["source"]).splitlines()
            if not line.lstrip().startswith(("%", "!"))
        ]
        ast.parse("\n".join(lines), filename=f"{path.name}:cell-{index}")


def extracted_functions(path, cell_index, names, namespace=None):
    notebook = load_notebook(path)
    tree = ast.parse("".join(notebook["cells"][cell_index]["source"]))
    constants = {"AI_SUMMARY", "AI_KEYWORDS", "AI_STATUS"}
    body = []
    for node in tree.body:
        if isinstance(node, ast.FunctionDef) and node.name in names:
            body.append(node)
        elif isinstance(node, ast.Assign) and any(
            isinstance(target, ast.Name) and target.id in constants for target in node.targets
        ):
            body.append(node)
    values = dict(namespace or {})
    exec(compile(ast.Module(body=body, type_ignores=[]), str(path), "exec"), values)
    return values


@pytest.mark.parametrize("path", NOTEBOOKS)
def test_notebook_json_python_and_clean_outputs(path):
    notebook = load_notebook(path)
    assert notebook["nbformat"] == 4
    ids = [cell["id"] for cell in notebook["cells"]]
    assert len(ids) == len(set(ids))
    assert all(not cell.get("outputs") for cell in notebook["cells"] if cell["cell_type"] == "code")
    parse_notebook_code(path)


@pytest.mark.parametrize("path", NOTEBOOKS)
def test_notebook_verifies_current_helper_bytes(path):
    setup = "".join(load_notebook(path)["cells"][2]["source"])
    expected = hashlib.sha256((ROOT / "zmo_common.py").read_bytes()).hexdigest()
    recorded = re.search(r'^HELPER_SHA256 = "([0-9a-f]{64})"$', setup, re.MULTILINE)
    commit = re.search(r'^HELPER_COMMIT = "([0-9a-f]{40})"$', setup, re.MULTILINE)
    assert recorded and recorded.group(1) == expected
    assert commit
    assert "raw.githubusercontent.com/fmadore/zmo-ai-pipelines/main" not in setup
    assert "actual_helper_sha256 != HELPER_SHA256" in setup


def test_notebook_calls_match_helper_function_keywords():
    helper_tree = ast.parse((ROOT / "zmo_common.py").read_text(encoding="utf-8"))
    signatures = {}
    for node in helper_tree.body:
        if not isinstance(node, ast.FunctionDef):
            continue
        arguments = node.args
        names = {
            argument.arg
            for argument in arguments.posonlyargs + arguments.args + arguments.kwonlyargs
        }
        signatures[node.name] = (names, arguments.kwarg is not None)

    for path in NOTEBOOKS:
        tree = ast.parse(
            "\n".join(
                line
                for line in notebook_code(load_notebook(path)).splitlines()
                if not line.lstrip().startswith(("%", "!"))
            )
        )
        for call in (node for node in ast.walk(tree) if isinstance(node, ast.Call)):
            if not (
                isinstance(call.func, ast.Attribute)
                and isinstance(call.func.value, ast.Name)
                and call.func.value.id == "zc"
                and call.func.attr in signatures
            ):
                continue
            allowed, has_kwargs = signatures[call.func.attr]
            if has_kwargs:
                continue
            for keyword in call.keywords:
                if keyword.arg is not None:
                    assert keyword.arg in allowed, (
                        f"{path.name} calls zc.{call.func.attr} with unsupported "
                        f"keyword {keyword.arg!r}"
                    )


def test_summary_workbook_functions_preserve_sheets_formulas_and_styles(tmp_path):
    path = ROOT / "Summary_Colab.ipynb"
    functions = extracted_functions(
        path,
        12,
        {"workbook_headers", "copy_header_style", "ensure_output_columns", "save_workbook_atomic"},
        {"copy": copy, "Path": Path, "tempfile": __import__("tempfile"), "os": __import__("os")},
    )
    source = tmp_path / "source.xlsx"
    output = tmp_path / "output.xlsx"
    workbook = Workbook()
    data = workbook.active
    data.title = "Data"
    other = workbook.create_sheet("Other")
    data.append(["OCR", "Formula"])
    data.append(["source", "=1+1"])
    data["A1"].fill = PatternFill(fill_type="solid", fgColor="00FF00")
    other["A1"] = "preserve"
    workbook.save(source)

    workbook = load_workbook(source, data_only=False)
    headers = functions["ensure_output_columns"](workbook["Data"], 1)
    workbook["Data"].cell(2, headers["AI Summary"], "summary")
    functions["save_workbook_atomic"](workbook, output)
    workbook.close()

    checked = load_workbook(output, data_only=False)
    assert checked.sheetnames == ["Data", "Other"]
    assert checked["Other"]["A1"].value == "preserve"
    assert checked["Data"]["B2"].value == "=1+1"
    assert checked["Data"]["A1"].fill.fgColor.rgb.endswith("00FF00")
    assert checked["Data"].cell(2, headers["AI Summary"]).value == "summary"
    checked.close()


def test_summary_schema_validation_rejects_too_few_keywords():
    functions = extracted_functions(
        ROOT / "Summary_Colab.ipynb",
        12,
        {"salvage_truncated_json", "parse_summary_response"},
        {"json": json, "re": re},
    )
    valid = json.dumps({"summary": "S", "keywords": ["a", "b", "c", "d", "e"]})
    invalid = json.dumps({"summary": "S", "keywords": ["a", "b"]})
    assert functions["parse_summary_response"](valid)[2] is True
    assert functions["parse_summary_response"](invalid)[2] is False


def test_audio_refuses_video_when_ffmpeg_is_unavailable(tmp_path):
    video = tmp_path / "interview.mp4"
    video.write_bytes(b"video")
    fake_zc = SimpleNamespace(have_ffmpeg=lambda: False)
    functions = extracted_functions(
        ROOT / "Audio_Transcription_Colab.ipynb",
        12,
        {"prepare_media"},
        {"Path": Path, "VIDEO_EXTENSIONS": {".mp4"}, "zc": fake_zc},
    )
    with pytest.raises(RuntimeError, match="soundtrack"):
        functions["prepare_media"](video)


def test_ocr_prompts_distinguish_diplomatic_and_normalized_outputs():
    code = notebook_code(load_notebook(ROOT / "OCR_HTR_Colab.ipynb"))
    assert "absolute precision" not in code.lower()
    assert "Preserve end-of-line hyphens exactly" in code
    assert "Remove a line-break hyphen only" in code
    assert "MEDIA_RESOLUTION_MEDIUM" in code
    assert "if is_pdf else types.MediaResolution.MEDIA_RESOLUTION_HIGH" in code


def test_summary_uses_openpyxl_and_batch_without_flattening():
    code = notebook_code(load_notebook(ROOT / "Summary_Colab.ipynb"))
    assert "pandas" not in code
    assert "read_excel" not in code
    assert "to_excel" not in code
    assert "client.batches.create" in code
    assert "AI Status" in code
    assert "minItems" in code and "maxItems" in code
